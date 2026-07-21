import openpyxl
import pytest

from loader.xlsx.openpyxl import Loader
from util.passthrough_layout import PassthroughLayout

pytestmark = pytest.mark.unit


@pytest.fixture
def xlsx_file(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "타이틀"
    ws1.append(["col1", "col2", "col3"])
    ws1.append([1, 2, 3])
    ws1.append([4, 5, 6])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["a", "b"])
    ws2.append([1, 2])
    ws2.append([None, None])
    ws2.append(["c", "d"])
    ws2.append([3, 4])

    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return str(path)


def test_get_layout_and_ocr(xlsx_file):
    loader = Loader(layout_config={"type": "detr"}, ocr_config={"type": "paddle"})
    assert isinstance(loader.layout, PassthroughLayout)
    assert loader.ocr is None


def test_one_table_item_per_sheet_and_blank_row_block(xlsx_file):
    loader = Loader()
    items = loader(xlsx_file)

    assert all(item["category"] == "table" for item in items)
    assert all(item["page"] == 1 for item in items)
    # Sheet1: 병합 타이틀 1개 표. Sheet2: 빈 행으로 나뉜 표 2개.
    assert len(items) == 3


def test_merged_title_row_becomes_caption(xlsx_file):
    loader = Loader()
    items = loader(xlsx_file)

    sheet1_item = items[0]
    assert "<caption>타이틀</caption>" in sheet1_item["text"]
    assert "<th>col1</th>" in sheet1_item["text"]
    assert "<td>1</td>" in sheet1_item["text"]


def test_blank_row_splits_sheet_into_separate_table_blocks(xlsx_file):
    loader = Loader()
    items = loader(xlsx_file)

    sheet2_items = items[1:]
    assert len(sheet2_items) == 2
    assert "<th>a</th>" in sheet2_items[0]["text"]
    assert "<th>c</th>" in sheet2_items[1]["text"]
